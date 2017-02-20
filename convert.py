import openpyxl, pprint

filepath = '/Users/coreyclip/Desktop/2015-raporti-vjetor-per-kontratat-e-nenshkruara-publike.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb.get_active_sheet()

#gather meta data
d = defaultdict(dict)
d['Data e p‘rgatitjes s‘ raportit'] = ws['G8'].value

d['Viti fiskal'] = ws['P8'].value

d['Emri zyrtar i Autoritetit Kontraktues'] = ws['I12'].value

d['Lloji i Autoritetit Kontraktues (zgjidh‘ne nj‘ren)'] = ws['I13'].value

d['Kodi buxhetor'] = ws['I14'].value

d['Adresa'] = ws['I15'].value

d['Kodi postar - Qyteti'] = ws['I16'].value

d['Tel. fiks/ Celulari /Faksi'] = ws['I17'].value

d['Emri i menaxherit t‘ prokurimit'] = ws['I18'].value

d['E-mail adresa'] = ws['I19'].value

d['Adresa e webit t‘ AK'] = ws['I20'].value

import csv 
outputfile = open('/Users/coreyclip/Desktop/meta_data.csv', 'w')
with open(outputfile, 'w') as csvfile:
    fieldnames = [
                'Data e p‘rgatitjes s‘ raportit', 
                'Viti fiskal',
                'Emri zyrtar i Autoritetit Kontraktues',
                'Lloji i Autoritetit Kontraktues (zgjidh‘ne nj‘ren)',
                'Kodi buxhetor',
                'Adresa',
                'Kodi postar - Qyteti',
                'Tel. fiks/ Celulari /Faksi',
                'Emri i menaxherit t‘ prokurimit',
                'E-mail adresa',
                'Adresa e webit t‘ AK',
                ]
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

    writer.writeheader()
    writer.writerow({
                    'Data e p‘rgatitjes s‘ raportit':ws['G8'].value, 
                    'Viti fiskal': ws['P8'].value,
                    'Emri zyrtar i Autoritetit Kontraktues': ws['I12'].value,
                    'Lloji i Autoritetit Kontraktues (zgjidh‘ne nj‘ren)':ws['I13'].value,
                    'Kodi buxhetor':ws['I14'].value,
                    'Adresa': ws['I15'].value,
                    'Kodi postar - Qyteti': ws['I16'].value,
                    'Tel. fiks/ Celulari /Faksi': ws['I17'].value,
                    'Emri i menaxherit t‘ prokurimit': ws['I18'].value,
                    'E-mail adresa': ws['I19'].value,
                    'Adresa e webit t‘ AK' : ws['I20'].value
                    })

#iterate through releases
#for row in ws.iter_rows(min_row=28, max_col=3, max_row=138):

